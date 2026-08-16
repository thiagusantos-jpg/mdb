"""F01 — importação da planilha de referência.

Sobre não usar pandas, como o PRD v1 previa: `pd.read_csv` infere tipos, e é
exatamente essa inferência que transforma o EAN 7891234567890 no float
7891234567890.0 e códigos longos em notação científica — a corrupção silenciosa
apontada na revisão (§3.1). O módulo `csv` da biblioteca padrão lê tudo como
texto, que é o que se quer aqui, e evita uma dependência de ~50 MB para ler
alguns milhares de linhas. `.xlsx` usa openpyxl, importado sob demanda.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

from .. import config
from ..normalize.precos import parse_preco_brl
from ..normalize.quantidades import extrair_do_titulo
from ..normalize.texto import normalizar_ean, normalizar_titulo, validar_ean


@dataclass
class LinhaPlanilha:
    numero: int
    ean: str | None
    nome: str
    preco_centavos: int
    marca: str | None = None
    categoria: str | None = None
    quantidade: float | None = None
    unidade: str | None = None
    ean_invalido: bool = False


@dataclass
class ResultadoLeitura:
    linhas: list[LinhaPlanilha] = field(default_factory=list)
    descartadas: list[tuple[int, str]] = field(default_factory=list)
    ean_sem_digito_valido: int = 0
    colunas_detectadas: dict[str, str] = field(default_factory=dict)

    @property
    def total_lidas(self) -> int:
        return len(self.linhas) + len(self.descartadas)


class ErroDePlanilha(RuntimeError):
    pass


# --------------------------------------------------------------------------
# Detecção de colunas
# --------------------------------------------------------------------------


def _chave(cabecalho: str) -> str:
    return normalizar_titulo(cabecalho).replace(" ", "_")


def _mapear_colunas(cabecalhos: list[str]) -> dict[str, str]:
    """Liga nomes de coluna da planilha aos campos que usamos.

    Aceita `Preco_Referencia` e `Preco`: o PRD v1 divergia de si mesmo nesse
    ponto (a spec dizia um, o código lia o outro), então as duas planilhas
    existem no mundo real.
    """
    normalizados = {_chave(c): c for c in cabecalhos if c}
    mapa: dict[str, str] = {}

    for campo, candidatos in (
        ("ean", config.COLUNAS_EAN),
        ("nome", config.COLUNAS_NOME),
        ("preco", config.COLUNAS_PRECO),
        ("marca", config.COLUNAS_MARCA),
        ("categoria", config.COLUNAS_CATEGORIA),
    ):
        for candidato in candidatos:
            if candidato in normalizados:
                mapa[campo] = normalizados[candidato]
                break

    faltando = [c for c in ("nome", "preco") if c not in mapa]
    if faltando:
        raise ErroDePlanilha(
            f"colunas obrigatórias ausentes: {', '.join(faltando)}. "
            f"Cabeçalhos encontrados: {', '.join(cabecalhos)}"
        )
    return mapa


# --------------------------------------------------------------------------
# Leitura bruta
# --------------------------------------------------------------------------


def _ler_csv(caminho: Path) -> Iterator[dict[str, str]]:
    # utf-8-sig descarta o BOM que o Excel escreve.
    with caminho.open("r", encoding="utf-8-sig", newline="") as arquivo:
        amostra = arquivo.read(8192)
        arquivo.seek(0)
        try:
            dialeto = csv.Sniffer().sniff(amostra, delimiters=",;\t")
        except csv.Error:
            dialeto = csv.excel  # planilha de uma coluna só, ou amostra curta
        yield from csv.DictReader(arquivo, dialect=dialeto)


def _ler_xlsx(caminho: Path) -> Iterator[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as erro:  # pragma: no cover - depende do ambiente
        raise ErroDePlanilha(
            "leitura de .xlsx requer openpyxl: pip install openpyxl "
            "(ou exporte a planilha como .csv)"
        ) from erro

    livro = load_workbook(filename=str(caminho), read_only=True, data_only=True)
    try:
        aba = livro.active
        linhas = aba.iter_rows(values_only=True)
        cabecalhos = [str(c) if c is not None else "" for c in next(linhas, ())]
        for valores in linhas:
            yield {
                cabecalho: valor
                for cabecalho, valor in zip(cabecalhos, valores)
                if cabecalho
            }
    finally:
        livro.close()


def ler_planilha(caminho: Path | str) -> ResultadoLeitura:
    """Lê e valida a planilha, sem tocar no banco.

    Linha inválida é descartada COM motivo registrado, nunca gravada como NaN.
    Na v1, um preço vazio virava float('nan'), era gravado, e depois sumia dos
    alertas comparando False em qualquer '<' — sem erro nenhum.
    """
    origem = Path(caminho)
    if not origem.exists():
        raise ErroDePlanilha(f"arquivo não encontrado: {origem}")

    sufixo = origem.suffix.lower()
    if sufixo in (".csv", ".txt", ".tsv"):
        registros = _ler_csv(origem)
    elif sufixo in (".xlsx", ".xlsm"):
        registros = _ler_xlsx(origem)
    else:
        raise ErroDePlanilha(f"formato não suportado: {sufixo} (use .csv ou .xlsx)")

    resultado = ResultadoLeitura()
    mapa: dict[str, str] | None = None

    for indice, bruto in enumerate(registros, start=2):  # linha 1 = cabeçalho
        if mapa is None:
            mapa = _mapear_colunas(list(bruto.keys()))
            resultado.colunas_detectadas = mapa

        nome = str(bruto.get(mapa["nome"]) or "").strip()
        if not nome:
            resultado.descartadas.append((indice, "nome vazio"))
            continue

        preco = parse_preco_brl(bruto.get(mapa["preco"]))
        if preco is None:
            resultado.descartadas.append(
                (indice, f"preço ilegível: {bruto.get(mapa['preco'])!r}")
            )
            continue
        if preco <= 0:
            resultado.descartadas.append((indice, f"preço não positivo: {preco}"))
            continue

        ean = normalizar_ean(bruto.get(mapa["ean"])) if "ean" in mapa else None
        ean_invalido = bool(ean) and not validar_ean(ean)
        if ean_invalido:
            resultado.ean_sem_digito_valido += 1

        quantidade, unidade, _ = extrair_do_titulo(nome)

        resultado.linhas.append(
            LinhaPlanilha(
                numero=indice,
                ean=ean,
                nome=nome,
                preco_centavos=preco,
                marca=(str(bruto.get(mapa["marca"])).strip() if "marca" in mapa and bruto.get(mapa["marca"]) else None),
                categoria=(str(bruto.get(mapa["categoria"])).strip() if "categoria" in mapa and bruto.get(mapa["categoria"]) else None),
                quantidade=quantidade,
                unidade=unidade,
                ean_invalido=ean_invalido,
            )
        )

    if mapa is None:
        raise ErroDePlanilha("planilha vazia (nenhuma linha de dados)")

    return resultado
